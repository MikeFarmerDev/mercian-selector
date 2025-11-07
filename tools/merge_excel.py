import os
import json
import csv
from datetime import datetime
from shutil import copyfile

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_PATH = os.path.join(BASE_DIR, "config", "sync_map.json")


def load_config(path=CONFIG_PATH):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def backup_excel(excel_path, backup_dir):
    os.makedirs(backup_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    base = os.path.basename(excel_path)
    backup_path = os.path.join(backup_dir, f"{os.path.splitext(base)[0]}_{ts}.xlsx")
    copyfile(excel_path, backup_path)
    return backup_path


def ensure_columns(ws, header_row, needed_cols):
    headers = {}
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[str(val).strip()] = col

    for col_name in needed_cols:
        if col_name not in headers:
            max_col += 1
            ws.cell(row=header_row, column=max_col).value = col_name
            headers[col_name] = max_col

    return headers


def load_shopify_csv(path):
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    return rows


def _norm(val):
    return str(val).strip().upper() if val is not None else ""


def main():
    config = load_config()

    excel_path = os.path.join(BASE_DIR, config["excel"]["path"])
    sheet_name = config["excel"]["sheet"]
    primary_key = config["excel"]["primary_key"]

    shopify_csv = os.path.join(BASE_DIR, config["shopify"]["flattened_csv"])
    report_csv = os.path.join(BASE_DIR, config["reporting"]["sync_report_csv"])
    coaching_csv = os.path.join(BASE_DIR, config["reporting"]["coaching_exceptions_csv"])
    backup_dir = os.path.join(BASE_DIR, config["reporting"]["backup_dir"])

    backup_path = backup_excel(excel_path, backup_dir)
    print(f"[sync] Backup created at {backup_path}")

    shopify_rows = load_shopify_csv(shopify_csv)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    create_if_missing = config["columns"]["create_if_missing"]
    never_overwrite = set(config["columns"]["never_overwrite"])
    allowed_updates = config["columns"]["allowed_updates"]
    vat_col = config["columns"]["vat_column"]
    default_vat = config["columns"].get("default_vat_rate", 0.2)

    # header = row 1
    needed_cols = list(create_if_missing) + [primary_key, "inventory_item_id"]
    headers = ensure_columns(ws, 1, needed_cols)

    # --- build Excel index (NORMALISED) ---
    excel_index = {}
    pk_col_idx = headers[primary_key]
    for row in range(2, ws.max_row + 1):
        pk_val = ws.cell(row=row, column=pk_col_idx).value
        if pk_val:
            excel_index[_norm(pk_val)] = row

    updated_rows = 0
    changed_cells = 0
    new_products = []
    coaching_exceptions = []

    # we'll collect Shopify codes (normalised) to later detect true missing
    shopify_codes_seen = set()

    for srow in shopify_rows:
        # get product code or sku from Shopify
        raw_pk = (
            srow.get("product_code")
            or srow.get("Product Code")
            or srow.get("sku")
            or srow.get("SKU")
        )
        if not raw_pk:
            # skip rows without any usable code
            continue

        pk = _norm(raw_pk)
        shopify_codes_seen.add(pk)

        # --- try to match Excel row ---
        excel_row = excel_index.get(pk)

        # if product_code didn't match, try SKU fallback explicitly
        if not excel_row:
            alt_sku = srow.get("sku") or srow.get("SKU")
            if alt_sku:
                alt_norm = _norm(alt_sku)
                if alt_norm in excel_index:
                    pk = alt_norm
                    excel_row = excel_index[alt_norm]
                    shopify_codes_seen.add(alt_norm)

        if excel_row:
            # existing product in Excel → update allowed fields
            # ensure VAT
            vat_cell = ws.cell(row=excel_row, column=headers[vat_col])
            if vat_cell.value is None:
                vat_cell.value = default_vat
            vat_val = float(vat_cell.value)

            for excel_col_name, source_expr in allowed_updates.items():
                if excel_col_name in never_overwrite:
                    continue
                target_idx = headers.get(excel_col_name)
                if not target_idx:
                    continue

                new_val = None
                if source_expr.startswith("shopify."):
                    field = source_expr.split(".", 1)[1]
                    new_val = srow.get(field, "")
                    if field == "active":
                        new_val = True if str(new_val).lower() in ("1", "true", "yes", "active") else False
                elif source_expr.startswith("calc:"):
                    # only current calc: price_ex_vat*(1+vat)
                    price_s = srow.get("price_ex_vat")
                    if price_s:
                        try:
                            price = float(price_s)
                            new_val = round(price * (1 + vat_val), 2)
                        except ValueError:
                            new_val = None

                if new_val is not None:
                    cell = ws.cell(row=excel_row, column=target_idx)
                    if cell.value != new_val:
                        cell.value = new_val
                        changed_cells += 1

                # ensure Shopify Active is written for existing rows
                shopify_status_col = headers.get("Shopify Status")
                shopify_active_col = headers.get("Shopify Active")

                status_val = str(srow.get("product_status", "")).strip().upper()
                is_active = (status_val == "ACTIVE")

                if shopify_status_col:
                    ws.cell(row=excel_row, column=shopify_status_col).value = status_val

                if shopify_active_col:
                    ws.cell(row=excel_row, column=shopify_active_col).value = is_active
            
            # --- extra Shopify → Excel fields from flattened CSV ---
            # 1) Image URL
            if "Image URL" in srow and "Image URL" in headers:
                cell = ws.cell(row=excel_row, column=headers["Image URL"])
                if cell.value in (None, "") and srow["Image URL"]:
                    cell.value = srow["Image URL"]
                    changed_cells += 1

            # 2) Product URL
            if "Product URL" in srow and "Product URL" in headers:
                cell = ws.cell(row=excel_row, column=headers["Product URL"])
                if cell.value in (None, "") and srow["Product URL"]:
                    cell.value = srow["Product URL"]
                    changed_cells += 1

            # 2a) Inventory Item ID (Excel-safe: force Text format)
            if "inventory_item_id" in srow and "inventory_item_id" in headers:
                inv_val = srow.get("inventory_item_id", "")
                # Strip ="...": keep the digits only
                if isinstance(inv_val, str) and inv_val.startswith('="') and inv_val.endswith('"'):
                    inv_val = inv_val[2:-1]
                inv_cell = ws.cell(row=excel_row, column=headers["inventory_item_id"], value=str(inv_val))
                inv_cell.number_format = "@"  # ensure Excel treats it as Text, not a number

            # 2b) Product Title
            if "title" in srow and "Title" in headers:
                cell = ws.cell(row=excel_row, column=headers["Title"])
                if cell.value in (None, "") and srow["title"]:
                    cell.value = srow["title"]
                    changed_cells += 1


            # 3) Description Narrative
            # map flattened column → Excel column
            src_desc = srow.get("global.description_tag", "")

            # find the actual Excel header even if it has spaces
            desc_header_name = None
            for hname in headers.keys():
                if hname and hname.strip() == "Description Narrative":
                    desc_header_name = hname
                    break

            if src_desc and desc_header_name:
                cell = ws.cell(row=excel_row, column=headers[desc_header_name])
                current_val = cell.value
                # treat None, "", and whitespace as blank
                if current_val is None or (isinstance(current_val, str) and current_val.strip() == ""):
                    cell.value = src_desc
                    changed_cells += 1


            if "Description Narrative" in srow and desc_header_name:
                cell = ws.cell(row=excel_row, column=headers[desc_header_name])
                current_val = cell.value
                # treat whitespace-only as blank
                is_blank = current_val is None or (isinstance(current_val, str) and current_val.strip() == "")
                if is_blank and srow["Description Narrative"]:
                    cell.value = srow["Description Narrative"]
                    changed_cells += 1

            # 3b) CSV Description Narrative → Excel Description
            if "Description Narrative" in srow and "Description" in headers:
                cell = ws.cell(row=excel_row, column=headers["Description"])
                if cell.value in (None, "") and srow["Description Narrative"]:
                    cell.value = srow["Description Narrative"]
                    changed_cells += 1

            # 4) price_ex_vat → Price (excluding VAT) + Full Price
            if "price_ex_vat" in srow and srow["price_ex_vat"]:
                try:
                    price_ex_vat = float(srow["price_ex_vat"])
                    # write Price (excluding VAT) if blank
                    if "Price (excluding VAT)" in headers:
                        price_cell = ws.cell(row=excel_row, column=headers["Price (excluding VAT)"])
                        if price_cell.value in (None, ""):
                            price_cell.value = price_ex_vat
                            changed_cells += 1
                    # now calc Full Price using row VAT
                    full_price = round(price_ex_vat * (1 + vat_val), 2)
                    if "Full Price" in headers:
                        full_cell = ws.cell(row=excel_row, column=headers["Full Price"])
                        if full_cell.value in (None, ""):
                            full_cell.value = full_price
                            changed_cells += 1
                except ValueError:
                    pass


            updated_rows += 1

        else:
            # new product from Shopify → add to Excel
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=headers[primary_key]).value = pk

            # set VAT
            ws.cell(row=new_row, column=headers[vat_col]).value = default_vat

            # fill allowed updates from Shopify
            for excel_col_name, source_expr in allowed_updates.items():
                if excel_col_name in never_overwrite:
                    continue
                target_idx = headers.get(excel_col_name)
                if not target_idx:
                    continue

                val = None
                if source_expr.startswith("shopify."):
                    field = source_expr.split(".", 1)[1]
                    val = srow.get(field, "")
                elif source_expr.startswith("calc:"):
                    price_s = srow.get("price_ex_vat")
                    if price_s:
                        try:
                            price = float(price_s)
                            val = round(price * (1 + default_vat), 2)
                        except ValueError:
                            val = None

                if val is not None:
                    ws.cell(row=new_row, column=target_idx).value = val

            # write Shopify Active for new rows too
            shopify_status_col = headers.get("Shopify Status")
            shopify_active_col = headers.get("Shopify Active")

            status_val = str(srow.get("product_status", "")).strip().upper()
            is_active = (status_val == "ACTIVE")

            if shopify_status_col:
                ws.cell(row=new_row, column=shopify_status_col).value = status_val

            if shopify_active_col:
                ws.cell(row=new_row, column=shopify_active_col).value = is_active

            new_products.append(pk)

            # CSV Description Narrative → Excel Description (for new rows)
            if "Description Narrative" in srow and "Description" in headers:
                cell = ws.cell(row=new_row, column=headers["Description"])
                if srow["Description Narrative"]:
                    cell.value = srow["Description Narrative"]

            # build coaching exception
            coaching_exceptions.append({
                "Product Code": pk,
                "Title": srow.get("title", ""),
                "Needs": "Drag Flicking, Aerial, Reverse Stick Hitting, Power, Touch and Control, 3D, Playing Level, Bow, Carbon, Length, Player Type"
            })

    # save workbook after all updates
    wb.save(excel_path)
    # === Inventory Join (config-driven path; sum across locations; 'notmatched' else) ===
    try:
        inv_csv_rel = config["inventory"]["flat_csv"]
    except KeyError:
        raise KeyError("sync_map.json missing key: inventory.flat_csv")

    inv_csv_path = os.path.join(BASE_DIR, inv_csv_rel)
    if os.path.exists(inv_csv_path):
        # load and aggregate inventory
        inv_rows = load_shopify_csv(inv_csv_path)

        def _norm_iid(x):
            s = str(x).strip() if x is not None else ""
            if s.startswith('="') and s.endswith('"'):
                s = s[2:-1]
            return s.strip()

        inv_agg = {}
        for r in inv_rows:
            iid = _norm_iid(r.get("inventory_item_id"))
            if not iid:
                continue
            try:
                avail = int(str(r.get("available", "0")).strip())
            except ValueError:
                avail = 0
            ts = str(r.get("updated_at", "")).strip()
            if iid not in inv_agg:
                inv_agg[iid] = {"available_sum": 0, "updated_at_latest": ""}
            inv_agg[iid]["available_sum"] += avail
            if ts and (inv_agg[iid]["updated_at_latest"] == "" or ts > inv_agg[iid]["updated_at_latest"]):
                inv_agg[iid]["updated_at_latest"] = ts

        # ensure destination columns exist
        headers = ensure_columns(ws, 1, ["inventory_item_id", "available", "updated_at_latest"])
        iid_col = headers["inventory_item_id"]
        avail_col = headers["available"]
        ts_col = headers["updated_at_latest"]

        notmatched = []

        for row in range(2, ws.max_row + 1):
            raw_iid = ws.cell(row=row, column=iid_col).value
            iid = _norm_iid(raw_iid) if raw_iid is not None else ""
            if iid and iid in inv_agg:
                ws.cell(row=row, column=avail_col).value = inv_agg[iid]["available_sum"]
                ws.cell(row=row, column=ts_col).value = inv_agg[iid]["updated_at_latest"]
            else:
                ws.cell(row=row, column=avail_col).value = "notmatched"
                ws.cell(row=row, column=ts_col).value = ""
                notmatched.append(iid)

        # persist after inventory join
        wb.save(excel_path)

        # append reporting
        os.makedirs(os.path.dirname(report_csv), exist_ok=True)
        with open(report_csv, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if os.stat(report_csv).st_size == 0:
                w.writerow(["timestamp", "stage", "total_matched", "total_notmatched"])
            ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
            w.writerow([ts_now, "inventory_join", len(inv_agg), len(notmatched)])

        # append coaching exceptions
        if notmatched:
            with open(coaching_csv, "a", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                if not os.path.exists(coaching_csv) or os.stat(coaching_csv).st_size == 0:
                    w.writerow(["timestamp", "reason", "inventory_item_id"])
                ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
                for iid in notmatched:
                    w.writerow([ts_now, "MISSING_INVENTORY", iid])


    # --- now compute "missing in Shopify" properly ---
    # We want: only rows that are Active in Excel AND (Shopify Active is true or not present) AND not seen in Shopify
    missing_in_shopify = []
    active_col_idx = headers.get("Active")  # may be None
    shopify_active_col_idx = headers.get("Shopify Active")  # may be None

    for norm_code, row_idx in excel_index.items():
        # if Shopify actually saw this code → not missing
        if norm_code in shopify_codes_seen:
            continue

        # read Excel Active
        is_excel_active = True
        if active_col_idx:
            excel_active_val = ws.cell(row=row_idx, column=active_col_idx).value
            is_excel_active = str(excel_active_val or "").strip().lower() in ("true", "1", "yes", "y")

        # read Shopify Active (from Excel column, not from Shopify CSV)
        is_shopify_active = True
        if shopify_active_col_idx:
            shopify_active_val = ws.cell(row=row_idx, column=shopify_active_col_idx).value
            if shopify_active_val is not None:
                is_shopify_active = str(shopify_active_val).strip().lower() in ("true", "1", "yes", "y")

        # only flag as missing if both are active
        if is_excel_active and is_shopify_active:
            missing_in_shopify.append(norm_code)

    # write main sync report
    os.makedirs(os.path.dirname(report_csv), exist_ok=True)
    with open(report_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["metric", "value"])
        w.writerow(["updated_rows", updated_rows])
        w.writerow(["new_in_shopify", len(new_products)])
        w.writerow(["missing_in_shopify", len(missing_in_shopify)])
        w.writerow(["changed_cells", changed_cells])
        w.writerow([])
        w.writerow(["NEW_IN_SHOPIFY_LIST"])
        for pk in new_products:
            w.writerow([pk])
        w.writerow([])
        w.writerow(["MISSING_IN_SHOPIFY_LIST"])
        for pk in sorted(missing_in_shopify):
            w.writerow([pk])

    # write coaching exceptions
    if coaching_exceptions:
        with open(coaching_csv, "w", newline="", encoding="utf-8") as f:
            fieldnames = ["Product Code", "Title", "Needs"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(coaching_exceptions)

    print(f"[sync] Done. Updated {updated_rows}, new {len(new_products)}, missing {len(missing_in_shopify)}, changed cells {changed_cells}")


if __name__ == "__main__":
    main()
